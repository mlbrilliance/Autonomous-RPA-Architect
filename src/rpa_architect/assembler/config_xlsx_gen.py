"""Config.xlsx generation for UiPath REFramework projects.

Creates the standard three-sheet configuration workbook with Settings,
Constants, and Assets sheets. Uses openpyxl for Excel generation with
formatting.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from rpa_architect.ir.schema import ProcessIR

logger = logging.getLogger(__name__)

# Standard REFramework Settings entries
_DEFAULT_SETTINGS: list[dict[str, str]] = [
    {
        "Name": "OrchestratorQueueName",
        "Value": "",
        "Description": "Name of the Orchestrator queue for transaction items.",
    },
    {
        "Name": "MaxRetryNumber",
        "Value": "3",
        "Description": "Maximum number of retries for each transaction item.",
    },
    {
        "Name": "logF_BusinessProcessName",
        "Value": "",
        "Description": "Business process name used in log fields.",
    },
    {
        "Name": "ExcelSettingsFilePath",
        "Value": "Data\\Config.xlsx",
        "Description": "Path to this configuration file.",
    },
    {
        "Name": "ShouldMarkJobAsFaulted",
        "Value": "False",
        "Description": "Whether to mark the job as faulted on system exception.",
    },
    {
        "Name": "MaxConsecutiveSystemExceptions",
        "Value": "3",
        "Description": "Max consecutive system exceptions before stopping.",
    },
]


def _apply_header_formatting(ws: Any, num_columns: int) -> None:
    """Apply bold formatting and auto-width to header row."""
    from openpyxl.styles import Font

    bold_font = Font(bold=True)

    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font

    # Auto-width columns
    for col_cells in ws.columns:
        max_length = 0
        column_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_config_xlsx(ir: ProcessIR, output_path: Path) -> None:
    """Generate the Config.xlsx workbook for a REFramework project.

    Creates three sheets:
    - **Settings**: Standard REFramework settings (queue name, retry count, etc.)
    - **Constants**: User-defined constants from the IR config dict
    - **Assets**: Credential and asset references from the IR

    Args:
        ir: The ProcessIR containing config, credentials, and process metadata.
        output_path: Full file path where Config.xlsx will be written.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Config.xlsx generation. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()

    # --- Settings sheet ---
    ws_settings = wb.active
    if ws_settings is None:
        ws_settings = wb.create_sheet("Settings")
    else:
        ws_settings.title = "Settings"

    ws_settings.append(["Name", "Value", "Description"])

    for entry in _DEFAULT_SETTINGS:
        name = entry["Name"]
        value = entry["Value"]
        description = entry["Description"]

        # Override defaults with IR values
        if name == "logF_BusinessProcessName":
            value = ir.process_name
        elif name == "MaxRetryNumber" and "MaxRetryNumber" in ir.config:
            value = ir.config["MaxRetryNumber"]
        elif name == "OrchestratorQueueName":
            # Find queue from credentials
            queues = [c for c in ir.credentials if c.type == "queue"]
            if queues:
                value = queues[0].name

        ws_settings.append([name, value, description])

    _apply_header_formatting(ws_settings, 3)

    # --- Constants sheet ---
    ws_constants = wb.create_sheet("Constants")
    ws_constants.append(["Name", "Value", "Description"])

    # Add user-defined constants from IR config
    # Exclude keys that are already in Settings
    settings_keys = {entry["Name"] for entry in _DEFAULT_SETTINGS}

    for key, value in sorted(ir.config.items()):
        if key not in settings_keys:
            ws_constants.append([key, value, ""])

    _apply_header_formatting(ws_constants, 3)

    # --- Assets sheet ---
    ws_assets = wb.create_sheet("Assets")
    ws_assets.append(["Name", "Type", "OrchestratorPath", "Description"])

    for cred in ir.credentials:
        ws_assets.append([
            cred.name,
            cred.type,
            cred.orchestrator_path or "",
            cred.description or "",
        ])

    _apply_header_formatting(ws_assets, 4)

    # Write file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    logger.info("Generated Config.xlsx at %s.", output_path)
