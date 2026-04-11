"""Tests for Config.xlsx generation."""

from __future__ import annotations

from pathlib import Path

import pytest

from rpa_architect.assembler.config_xlsx_gen import generate_config_xlsx
from rpa_architect.ir.schema import ProcessIR


class TestGenerateConfigXlsx:
    """Test that Config.xlsx is created with correct sheets."""

    def test_generate_config_xlsx(self, sample_ir: ProcessIR, tmp_output_dir: Path) -> None:
        output_path = tmp_output_dir / "Config.xlsx"
        generate_config_xlsx(sample_ir, output_path)

        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Verify sheets via openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))
        sheet_names = wb.sheetnames
        assert "Settings" in sheet_names
        assert "Constants" in sheet_names
        assert "Assets" in sheet_names
        wb.close()


class TestSettingsSheet:
    """Test that the Settings sheet has expected rows."""

    def test_config_xlsx_settings_sheet(self, sample_ir: ProcessIR, tmp_output_dir: Path) -> None:
        output_path = tmp_output_dir / "Config.xlsx"
        generate_config_xlsx(sample_ir, output_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))
        ws = wb["Settings"]

        # Header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["Name", "Value", "Description"]

        # Collect all setting names
        setting_names = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                setting_names.append(name)

        # Standard REFramework settings should be present
        assert "MaxRetryNumber" in setting_names
        assert "logF_BusinessProcessName" in setting_names
        assert "OrchestratorQueueName" in setting_names
        assert "ShouldMarkJobAsFaulted" in setting_names

        # logF_BusinessProcessName should be set to the process name
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "logF_BusinessProcessName":
                assert ws.cell(row=row, column=2).value == sample_ir.process_name
                break

        wb.close()


class TestConstantsSheet:
    """Test that Constants sheet contains IR config entries (non-Settings keys)."""

    def test_config_xlsx_constants_sheet(self, sample_ir: ProcessIR, tmp_output_dir: Path) -> None:
        output_path = tmp_output_dir / "Config.xlsx"
        generate_config_xlsx(sample_ir, output_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))
        ws = wb["Constants"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["Name", "Value", "Description"]

        constant_names = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                constant_names.append(name)

        # Config entries NOT in default Settings should appear here
        assert "InvoicePortalUrl" in constant_names
        assert "HighValueThreshold" in constant_names

        # MaxRetryNumber is a Settings key, should NOT be in Constants
        assert "MaxRetryNumber" not in constant_names

        wb.close()


class TestAssetsSheet:
    """Test that Assets sheet has credential entries from the IR."""

    def test_config_xlsx_assets_sheet(self, sample_ir: ProcessIR, tmp_output_dir: Path) -> None:
        output_path = tmp_output_dir / "Config.xlsx"
        generate_config_xlsx(sample_ir, output_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))
        ws = wb["Assets"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["Name", "Type", "OrchestratorPath", "Description"]

        # Check that our credential is listed
        asset_names = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                asset_names.append(name)

        assert "InvoicePortal_Cred" in asset_names

        # Check the credential type
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "InvoicePortal_Cred":
                assert ws.cell(row=row, column=2).value == "credential"
                assert ws.cell(row=row, column=3).value == "Production/InvoicePortal_ServiceAccount"
                break

        wb.close()
